# -*- coding: utf-8 -*-
"""יצירת קובץ אקסל מעוצב (RTL) עם כל המכרזים מסודרים לפי חודשים."""
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_FILE = os.path.join(os.path.dirname(__file__), "tenders.xlsx")

HEB_MONTHS = ["", "ינואר", "פברואר", "מרץ", "אפריל", "מאי", "יוני",
              "יולי", "אוגוסט", "ספטמבר", "אוקטובר", "נובמבר", "דצמבר"]

# כותרות העמודות
COLUMNS = [
    ("מכרז", 46),
    ("סוג המכרז", 15),
    ("מיקום", 16),
    ("שטח (מ״ר)", 13),
    ("תאריך פרסום", 14),
    ("מועד אחרון להגשה", 17),
    ("מפרסם", 20),
    ("פרטי קשר", 30),
    ("קישור", 14),
]

GREEN = "2E8B57"
GOLD = "C9971C"
LIGHT = "F3F7F3"
WHITE = "FFFFFF"

_thin = Side(style="thin", color="DDDDDD")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _month_key(t):
    """(year, month) לפי תאריך פרסום, ואם אין — לפי תאריך הגילוי."""
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", t.open_date or "")
    if m:
        return int(m.group(3)), int(m.group(2))
    if t.first_seen:
        try:
            y, mo = t.first_seen[:7].split("-")
            return int(y), int(mo)
        except ValueError:
            pass
    return (0, 0)


def _contact(t):
    return " · ".join(x for x in [t.contact_name, t.contact_phone, t.contact_email] if x)


def build(tenders):
    wb = Workbook()
    ws = wb.active
    ws.title = "מכרזים חקלאיים"
    ws.sheet_view.rightToLeft = True

    ncols = len(COLUMNS)
    for i, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # כותרת עליונה
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    top = ws.cell(row=1, column=1, value="🌾  מכרזים חקלאיים · משק אנג׳ל · אזור עמק יזרעאל")
    top.font = Font(bold=True, size=15, color=WHITE)
    top.fill = PatternFill("solid", fgColor=GREEN)
    top.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[1].height = 28

    # קיבוץ לפי חודש, מהחדש לישן
    groups = {}
    for t in tenders:
        groups.setdefault(_month_key(t), []).append(t)

    row = 3
    for key in sorted(groups, reverse=True):
        year, month = key
        label = f"{HEB_MONTHS[month]} {year}" if month else "ללא תאריך"

        # רצועת חודש
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        band = ws.cell(row=row, column=1, value=f"📅  {label}   ({len(groups[key])})")
        band.font = Font(bold=True, size=12, color=WHITE)
        band.fill = PatternFill("solid", fgColor=GOLD)
        band.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        # שורת כותרות עמודות
        for c, (name, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row, column=c, value=name)
            cell.font = Font(bold=True, color=WHITE, size=11)
            cell.fill = PatternFill("solid", fgColor=GREEN)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
        row += 1

        # שורות מכרזים
        items = sorted(groups[key], key=lambda t: t.open_date or "", reverse=True)
        for idx, t in enumerate(items):
            values = [
                t.title, t.ttype, t.location,
                t.area, t.open_date, t.close_date,
                t.publisher, _contact(t), "פתיחה ↗",
            ]
            fill = LIGHT if idx % 2 else WHITE
            for c, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=(c == 1))
                cell.font = Font(size=11)
            # קישור לחיץ בעמודת הכותרת ובעמודת "קישור"
            if t.url:
                for lc in (1, ncols):
                    link_cell = ws.cell(row=row, column=lc)
                    link_cell.hyperlink = t.url
                    link_cell.font = Font(size=11, color="1155CC", underline="single")
            # מועד אחרון בולט באדום
            if t.close_date:
                ws.cell(row=row, column=6).font = Font(size=11, bold=True, color="B23B2E")
            row += 1
        row += 1  # רווח בין חודשים

    wb.save(OUT_FILE)
    print(f"  📊 נוצר {os.path.basename(OUT_FILE)} ({len(tenders)} מכרזים)")
    return OUT_FILE
